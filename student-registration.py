
# Importojmë bibliotekat e nevojshme për GUI dhe operime të tjera
from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

# Definojmë ngjyrat për dizajnin
background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

# Krijojmë pjesen kryesore të aplikacionit
root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+210")
root.config(bg=background)

# Kontrollojmë nëse ekziston file 'Student_data.xlsx', nëse jo, e krijojmë atë me titujt përkatës
file = pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Date of Birth"
    sheet['F1'] = "Date of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skills"
    sheet['I1'] = "Father Name"
    sheet['J1'] = "Mother Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"
    file.save('Student_data.xlsx')

# Funksioni për daljen nga aplikacioni
def Exit():
    root.destroy()

# Funksioni për shfaqjen e imazhit të ngarkuar
def showImage():
    global filename
    global img
    filename = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select image file",
        filetype=(("JPG File", "*.jpg"),
                                                                             ("PNG File", "*.png"),
                                                                             ("All Files", "*.txt")))
    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

# Funksioni për krijimin e numrit të regjistrimit automatik
def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value

    
    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set("1")

# Funksioni për të pastruar të dhënat
def Clear():
    global img
    Name.set('')
    DateofBirth.set('')
    Religion.set('')
    Skills.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Class.set("Select Class")

    registration_no()
    saveButton.config(state='normal')
    img1 = PhotoImage(file='fotot/upload photo.png')
    lbl.config(image=img1)
    lbl.image = img1
    img = ""

# Funksioni për të ruajtur të dhënat
def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    
    try:
        G1 = gender  # Pasi që nuk ka përzgjedhje, do të tregojë gabim këtu
    except:
        messagebox.showerror("error", "Select Gender!")
        return  # Dalim nga funksioni nëse nuk është zgjedhur gjinia

    D2 = DateofBirth.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skills.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    # print(R1)
    # print(N1)
    # print(C1)
    # print(G1)
    # print(D2)
    # print(D1)
    # print(Re1)
    # print(S1)
    # print(fathername)
    # print(mothername)
    # print(F1)
    # print(M1)


    # Kontrollojmë nëse të dhënat janë të plota
    if N1 == "" or C1 == "Select Class" or D2 == "" or Re1 == "" or S1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "":
        messagebox.showerror("error", "Few Data is missing!")
    else:
        file = openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=C1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)  # Gjinia tani është e definuar
        sheet.cell(column=5, row=sheet.max_row, value=D2)
        sheet.cell(column=6, row=sheet.max_row, value=D1)
        sheet.cell(column=7, row=sheet.max_row, value=Re1)
        sheet.cell(column=8, row=sheet.max_row, value=S1)
        sheet.cell(column=9, row=sheet.max_row, value=fathername)
        sheet.cell(column=10, row=sheet.max_row, value=mothername)
        sheet.cell(column=11, row=sheet.max_row, value=F1)
        sheet.cell(column=12, row=sheet.max_row, value=M1)

        file.save(r'Student_data.xlsx')

        try:
            img.save("Student Images/" + str(R1) + ".jpg")
        except:
            messagebox.showinfo("info", "Profile is not available!")

        messagebox.showinfo("info", "Successfully data entered!")
        Clear()  # Pastrimi i fushave pas ruajtjes
        registration_no()  # Krijimi i numrit të ri të regjistrimit

###Funksioni per butoni  Search###
def search():
    
    text = Search.get() #taking input from entry box

    Clear() # clears all the date already availible in entry box and other
    saveButton.config(state='disable')# after clickin on search  save button will disable so that no one can click on it

    file= openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name= row[0]
#            print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number=str(name)[15:-1]

            # print(reg_no_position)
            # print(reg_number)
    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid", "Invalid registration number!!")

    #reg_no_position showing like A2, A3, A4, ... An
    #but reg_number just showng number afte A2 like 2, 3, ...,n

    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value

    # print(x1)
    # print(x2)
    # print(x3)
    # print(x4)
    # print(x5)
    # print(x6)
    # print(x7)
    # print(x8)
    # print(x9)
    # print(x10)
    # print(x11)
    # print(x12)

    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)

    if x4=='Female':
        R2.select()
    else:
        R1.select()

    DateofBirth.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skills.set(x8)
    F_Name.set(x9)
    M_Name.set(x10)
    Father_Occupation.set(x11)
    Mother_Occupation.set(x12)

    img= (Image.open("Student Images/" +str(x1)+".jpg"))# done this to take image name same as registration no
    resized_image=img.resize((190,190))
    photo2 =ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

######Update#####
def Update():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()

    selection()
    G1=gender

    D2 = DateofBirth.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skills.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet.file.active

    for row in sheet.rows:
        if row[0] == R1:
            name= row[0]
            print(str(name))
            reg_no_positions= str(name)[14:-1]
            reg_number= str(name)[15:-1]

            print(reg_number)
    
    #sheet.cell(column=1,row=int(reg_number), value=R1)
    sheet.cell(column=2,row=int(reg_number), value=N1)
    sheet.cell(column=3,row=int(reg_number), value=C1)
    sheet.cell(column=4,row=int(reg_number), value=G1)
    sheet.cell(column=5,row=int(reg_number), value=D2)
    sheet.cell(column=6,row=int(reg_number), value=D1)
    sheet.cell(column=7,row=int(reg_number), value=Re1)
    sheet.cell(column=8,row=int(reg_number), value=S1)
    sheet.cell(column=9,row=int(reg_number), value=fathername)
    sheet.cell(column=10,row=int(reg_number), value=mothername)
    sheet.cell(column=11,row=int(reg_number), value=F1)
    sheet.cell(column=12,row=int(reg_number), value=M1)

    file.save(r'Student_data.xlsx')

    try:
        img.save("Student Images/"+str(R1)+".jpg")
    except:
        pass

    messagebox.showinfo("Update", "Update Successfully!")

    Clear() 



# Funksioni për përzgjedhjen e gjinisë
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"

# Labelat e titujve dhe detajeve të aplikacionit
Label(root, text="Email: filanfisteku@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=3, bg="#c36466", fg="#fff", font="arial 20 bold").pack(side=TOP, fill=X)

# Kuti kërkimit search()
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font="arial 20").place(x=820, y=70)
imageicon3 = PhotoImage(file="fotot/search.png")
Srch = Button(root, text="Search", compound=LEFT, image=imageicon3, width=123, bg='#68ddfa', font="arial 13 bold", command=search)
Srch.place(x=1060, y=66)

imageicon4 = PhotoImage(file="fotot/update.png")
Update_button = Button(root, image=imageicon4, bg="#c36464",command=Update)
Update_button.place(x=110, y=64)

# Labelat për regjistrimin dhe datën
Label(root, text="Registration No:", font="arial 13", fg=framebg, bg=background).place(x=30, y=150)
Label(root, text="Date:", font="arial 13", fg=framebg, bg=background).place(x=500, y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)

registration_no()  # E thërrasim për të krijuar një numër regjistrimi

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)

Date.set(d1)

# Detajet e studentit
obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Full Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Date of Birth:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Gender:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)
Label(obj, text="Class:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Religion:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Skills:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

# Input për emrin, datën e lindjes dhe të tjera
Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160, y=50)

DateofBirth = StringVar()
Dob_entry = Entry(obj, textvariable=DateofBirth, width=20, font="arial 10")
Dob_entry.place(x=160, y=100)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=150, y=150)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=200, y=150)

Religion = StringVar()
religion_entry = Entry(obj, textvariable=Religion, width=20, font="arial 10")
religion_entry.place(x=630, y=100)

Skills = StringVar()
skills_entry = Entry(obj, textvariable=Skills, width=20, font="arial 10")
skills_entry.place(x=630, y=150)

# Kuti për klasën
Class = Combobox(obj, values=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'], font="Roboto 10", width=12, state="r")
Class.place(x=630, y=50)
Class.set("Select Class")

# Detajet e prindërve
obj2 = LabelFrame(root, text="Parent's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Father's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj2, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)

F_Name = StringVar()
f_entry = Entry(obj2, textvariable=F_Name, width=20, font="arial 10")
f_entry.place(x=160, y=50)

Father_Occupation = StringVar()
FO_entry = Entry(obj2, textvariable=Father_Occupation, width=20, font="arial 10")
FO_entry.place(x=160, y=100)

Label(obj2, text="Mother's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj2, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)

M_Name = StringVar()
m_entry = Entry(obj2, textvariable=M_Name, width=20, font="arial 10")
m_entry.place(x=630, y=50)

Mother_Occupation = StringVar()
MO_entry = Entry(obj2, textvariable=Mother_Occupation, width=20, font="arial 10")
MO_entry.place(x=630, y=100)

# Pjesa për ngarkimin e imazhit
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img = PhotoImage(file="fotot/upload photo.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

# Butonat për ngarkimin, ruajtjen dhe pastrimin
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showImage).place(x=1000, y=370)
saveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=Save)
saveButton.place(x=1000, y=450)
Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=Clear).place(x=1000, y=530)
Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="red", command=Exit).place(x=1000, y=610)

# Dëgjojmë ngjarjet e aplikacionit
root.mainloop()
